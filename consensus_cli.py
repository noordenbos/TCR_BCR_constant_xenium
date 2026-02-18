#!/usr/bin/env python3
"""Consensus finder for IMGT exon FASTA sequences."""

from __future__ import annotations

import argparse
import csv
import json
import os
import platform
import re
import shlex
import subprocess
import sys
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple


DNA_CHARS = set("acgtn")


@dataclass
class ExonResult:
    gene: str
    exon: str
    consensus: str
    subtype_rows: Dict[str, str]
    relation_rows: Dict[str, str]
    mismatch_counts: Dict[str, int]
    shared_start: Optional[int]
    shared_end: Optional[int]


@dataclass
class JunctionAnnotation:
    allele: str
    base_gene: str
    original_region: str
    split_source: str
    exon_chain: str
    junction_boundaries_1based: str
    note: str


@dataclass
class GeneConsensusSummary:
    gene: str
    source: str
    exon_chain: str
    exon_relative_coords: str
    exon_lengths: str
    consensus: str


@dataclass
class SharedElement:
    gene_a: str
    gene_b: str
    start_a_1based: int
    end_a_1based: int
    start_b_1based: int
    end_b_1based: int
    length: int
    mismatch_count: int
    seq_a: str
    seq_b: str


@dataclass
class PairConsensusResult:
    gene_a: str
    gene_b: str
    aligned_a: str
    aligned_b: str
    mirrored: str
    consensus: str


@dataclass
class RunMetadata:
    input_path: str
    output_xlsx: str
    output_fasta: str
    cwd: str
    cli_call: str
    cli_call_effective: str
    timestamp_utc: str
    python_version: str
    platform: str
    git_commit: str
    git_branch: str
    git_remote_origin: str
    git_remotes: str
    exon_reference_fasta: str
    compare_pairs: str
    family_mappings: str


def parse_gene_name(name: str) -> str:
    if "*" not in name:
        return name
    return name.split("*", 1)[0]


def clean_fasta_line(raw: str) -> str:
    line = raw.strip()
    if not line:
        return ""

    if ">" in line:
        line = line[line.index(">") :]
        return line.rstrip("\\").strip()

    if line.startswith("\\") or line.startswith("{") or line == "}":
        return ""

    return line.rstrip("\\").strip()


def parse_imgt_header(header: str) -> Tuple[str, str, str, int, int]:
    fields = header[1:].split("|")
    if len(fields) < 5:
        raise ValueError(f"Invalid IMGT header (expected >=5 fields): {header}")
    accession = fields[0].strip()
    allele = fields[1].strip()
    exon = fields[4].strip() or "UNKNOWN"
    pos_field = fields[5].strip() if len(fields) > 5 else ""
    trim_5p = 1 if pos_field.lower().startswith("a,") else 0
    trim_3p = 0
    return accession, allele, exon, trim_5p, trim_3p


def add_unique_subtype(group: Dict[str, str], key: str, seq: str) -> None:
    if key not in group:
        group[key] = seq
        return
    if group[key] == seq:
        return
    i = 2
    while f"{key}#{i}" in group:
        i += 1
    group[f"{key}#{i}"] = seq


def normalize_dna(seq: str) -> str:
    seq = (seq or "").strip().lower()
    return "".join(ch for ch in seq if ch in DNA_CHARS)


def parse_coordinate_part(part: str) -> Tuple[bool, int]:
    m = re.match(r"^([aA],)?\s*(\d+)\.\.(\d+)$", part.strip())
    if not m:
        raise ValueError(f"Invalid coordinate segment: {part}")
    has_added_5p = bool(m.group(1))
    start = int(m.group(2))
    end = int(m.group(3))
    genomic_len = abs(end - start) + 1
    return has_added_5p, genomic_len


def sort_exon_labels(labels: Iterable[str]) -> List[str]:
    def key(lbl: str) -> Tuple[int, int, str]:
        m = re.match(r"^EX(\d+)([A-Z]*)$", lbl)
        if not m:
            return (999, 999, lbl)
        num = int(m.group(1))
        suffix = m.group(2)
        suffix_rank = {"T": 0, "R": 1, "": 2}.get(suffix, 10)
        return (num, suffix_rank, suffix)

    return sorted(labels, key=key)


def exon_family_label(label: str) -> str:
    return label


def sort_exon_family_labels(labels: Iterable[str]) -> List[str]:
    def key(lbl: str) -> Tuple[int, int, str]:
        m = re.match(r"^EX(\d+)", lbl)
        if not m:
            return (999, 999, lbl)
        return (int(m.group(1)), 1, lbl)

    return sorted(labels, key=key)


def merge_variant_consensus(seqs: Sequence[str]) -> str:
    if not seqs:
        return ""
    if len(seqs) == 1:
        return seqs[0]
    if len({len(s) for s in seqs}) != 1:
        aligned = build_alignment({f"s{i}": s for i, s in enumerate(seqs, start=1)})
        rows = [aligned[k] for k in sorted(aligned.keys())]
        out: List[str] = []
        for i in range(len(rows[0])):
            col = [r[i] for r in rows if r[i] != "-"]
            if not col:
                continue
            if all(ch == col[0] for ch in col):
                out.append(col[0])
            else:
                out.append("X")
        return "".join(out)

    length = len(seqs[0])
    out: List[str] = []
    for i in range(length):
        col = [s[i] for s in seqs]
        if all(ch == col[0] for ch in col):
            out.append(col[0])
        elif all(ch == "x" for ch in col):
            out.append("x")
        else:
            out.append("X")
    return "".join(out)


def infer_segments_from_exon_lookup(
    allele: str,
    seq: str,
    exon_lookup: Dict[Tuple[str, str], str],
) -> Optional[List[Tuple[str, str]]]:
    exon_labels = sort_exon_labels(lbl for (a, lbl) in exon_lookup.keys() if a == allele and lbl.startswith("EX"))
    if not exon_labels:
        return None

    def run(candidate_seq: str) -> Optional[List[Tuple[str, str]]]:
        pos = 0
        out: List[Tuple[str, str]] = []
        for lbl in exon_labels:
            ref = exon_lookup[(allele, lbl)]
            if candidate_seq.startswith(ref, pos):
                out.append((lbl, ref))
                pos += len(ref)
                continue
            # Common IMGT edge case: leading ambiguous nt removed in curated combined sequence.
            if ref.startswith("n") and candidate_seq.startswith(ref[1:], pos):
                seg = ref[1:]
                out.append((lbl, seg))
                pos += len(seg)
                continue

        if pos == len(candidate_seq) and len(out) >= 2:
            return out
        return None

    direct = run(seq)
    if direct:
        return direct
    if len(seq) > 1:
        shifted = run(seq[1:])
        if shifted:
            return shifted
    return None


def junction_boundaries_1based(segments: Sequence[Tuple[str, str]]) -> List[int]:
    out: List[int] = []
    pos = 0
    for _, seg in segments[:-1]:
        pos += len(seg)
        out.append(pos)
    return out


def split_compiled_row(
    gene: str,
    coordinate: str,
    region: str,
    seq: str,
    exon_lookup: Optional[Dict[Tuple[str, str], str]] = None,
) -> Tuple[List[Tuple[str, str]], str, str]:
    seq = normalize_dna(seq)
    if not seq:
        return [], "empty", ""

    region_parts = [p.strip() for p in region.split("+") if p.strip()]
    coord_parts = [p.strip() for p in coordinate.split("+") if p.strip()] if coordinate else []

    if not region_parts:
        raise ValueError(f"{gene}: empty region field")

    if len(region_parts) == 1 and len(coord_parts) <= 1:
        if exon_lookup:
            inferred = infer_segments_from_exon_lookup(gene, seq, exon_lookup)
            if inferred:
                return inferred, "inferred_from_exons", "single-region row inferred from paired exon FASTA"
        return [(region_parts[0], seq)], "single_region_unsplit", "no exon inference available"

    if len(region_parts) > 1 and not coord_parts:
        if not exon_lookup:
            raise ValueError(f"{gene}: missing coordinates and no exon reference provided")
        ref_segments: List[str] = []
        for exon_label in region_parts:
            ref = exon_lookup.get((gene, exon_label))
            if not ref:
                raise ValueError(f"{gene}: missing reference exon sequence for {exon_label}")
            ref_segments.append(ref)

        split_lens = [len(s) for s in ref_segments]
        total_len = sum(split_lens)
        if total_len != len(seq):
            # Curated combined sequences may remove leading ambiguous 'n' base(s).
            delta = total_len - len(seq)
            if delta > 0:
                for i in range(len(ref_segments)):
                    if delta == 0:
                        break
                    if ref_segments[i].startswith("n") and split_lens[i] > 1:
                        split_lens[i] -= 1
                        ref_segments[i] = ref_segments[i][1:]
                        delta -= 1
            if sum(split_lens) != len(seq):
                raise ValueError(
                    f"{gene}: missing-coordinate split failed, combined fasta length {len(seq)} "
                    f"!= sum(reference exon lengths) {sum(split_lens)}"
                )
        out: List[Tuple[str, str]] = []
        pos = 0
        for exon_label, seg_len in zip(region_parts, split_lens):
            nxt = pos + seg_len
            out.append((exon_label, seq[pos:nxt]))
            pos = nxt
        return out, "region_list_no_coord", "split by region labels with paired exon FASTA lengths"

    if len(coord_parts) != len(region_parts):
        raise ValueError(
            f"{gene}: region/coordinate segment mismatch ({len(region_parts)} regions vs {len(coord_parts)} coordinates)"
        )

    coord_meta = [parse_coordinate_part(p) for p in coord_parts]
    genomic_lens = [glen for _, glen in coord_meta]
    with_added_lens = [glen + (1 if has_a else 0) for has_a, glen in coord_meta]

    seq_len = len(seq)
    base_total = sum(genomic_lens)
    with_added_total = sum(with_added_lens)

    if seq_len == with_added_total:
        split_lens = with_added_lens
        trim_5p_flags = [has_a for has_a, _ in coord_meta]
    elif seq_len == base_total:
        split_lens = genomic_lens
        trim_5p_flags = [False] * len(coord_meta)
    else:
        raise ValueError(
            f"{gene}: combined fasta length {seq_len} does not match coordinate-derived totals "
            f"(genomic={base_total}, with_a={with_added_total})"
        )

    segments_raw: List[str] = []
    pos = 0
    for seg_len in split_lens:
        nxt = pos + seg_len
        if nxt > seq_len:
            raise ValueError(f"{gene}: split overflow at position {pos} for segment length {seg_len}")
        segments_raw.append(seq[pos:nxt])
        pos = nxt
    if pos != seq_len:
        raise ValueError(f"{gene}: split underflow ({pos} != {seq_len})")

    segments: List[str] = []
    for seg_raw, trim_5p, genomic_len in zip(segments_raw, trim_5p_flags, genomic_lens):
        seg = seg_raw[1:] if trim_5p and seg_raw else seg_raw
        if len(seg) != genomic_len:
            raise ValueError(
                f"{gene}: segment length mismatch after trimming ({len(seg)} != expected {genomic_len})"
            )
        segments.append(seg)

    # Assert junction boundaries are represented in the combined fasta split.
    if len(segments) > 1:
        cum = 0
        for seg in segments[:-1]:
            cum += len(seg)
            if cum <= 0 or cum >= sum(len(s) for s in segments):
                raise ValueError(f"{gene}: invalid junction boundary {cum}")

    return list(zip(region_parts, segments)), "coordinate_split", ""


def read_compiled_csv_sequences(
    csv_path: Path,
    exon_lookup: Optional[Dict[Tuple[str, str], str]] = None,
) -> Tuple[Dict[Tuple[str, str], Dict[str, str]], List[JunctionAnnotation]]:
    grouped: Dict[Tuple[str, str], Dict[str, str]] = defaultdict(dict)
    annotations: List[JunctionAnnotation] = []
    with csv_path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        if not reader.fieldnames:
            raise ValueError("CSV has no header.")

        fields = {k.strip().lower(): k for k in reader.fieldnames}
        gene_key = fields.get("gene")
        coord_key = fields.get("coordinates") or fields.get("coordinate")
        region_key = fields.get("region")
        fasta_key = fields.get("fasta")
        if not gene_key or not region_key or not fasta_key:
            raise ValueError("CSV must contain columns: gene, region, fasta, and optionally coordinate(s).")

        for row in reader:
            full_gene = (row.get(gene_key) or "").strip()
            region = (row.get(region_key) or "").strip()
            seq = row.get(fasta_key) or ""
            coord = (row.get(coord_key) or "").strip() if coord_key else ""
            if not full_gene or not region or not seq:
                continue

            exon_segments, split_source, note = split_compiled_row(full_gene, coord, region, seq, exon_lookup=exon_lookup)
            base_gene = parse_gene_name(full_gene)
            for exon_label, exon_seq in exon_segments:
                add_unique_subtype(grouped[(base_gene, exon_label)], full_gene, exon_seq)

            boundaries = junction_boundaries_1based(exon_segments)
            annotations.append(
                JunctionAnnotation(
                    allele=full_gene,
                    base_gene=base_gene,
                    original_region=region,
                    split_source=split_source,
                    exon_chain="+".join(lbl for lbl, _ in exon_segments),
                    junction_boundaries_1based="+".join(str(v) for v in boundaries),
                    note=note,
                )
            )

    if not grouped:
        raise ValueError("No valid rows found in compiled CSV.")
    return grouped, annotations


def read_imgt_fasta_sequences(fasta_path: Path) -> Dict[Tuple[str, str], Dict[str, str]]:
    grouped: Dict[Tuple[str, str], Dict[str, str]] = defaultdict(dict)

    current_header: Optional[str] = None
    current_seq: List[str] = []

    def flush_record() -> None:
        nonlocal current_header, current_seq
        if not current_header:
            return
        seq = "".join(current_seq).lower()
        seq = "".join(ch for ch in seq if ch in DNA_CHARS)
        if not seq:
            current_header = None
            current_seq = []
            return

        accession, allele, exon, added_5p, added_3p = parse_imgt_header(current_header)
        if added_5p > 0:
            seq = seq[added_5p:] if added_5p < len(seq) else ""
        if added_3p > 0:
            seq = seq[:-added_3p] if added_3p < len(seq) else ""
        if not seq:
            current_header = None
            current_seq = []
            return
        base_gene = parse_gene_name(allele)
        subtype_key = allele if allele else accession
        add_unique_subtype(grouped[(base_gene, exon)], subtype_key, seq)
        current_header = None
        current_seq = []

    with fasta_path.open("r", encoding="utf-8", errors="ignore") as handle:
        for raw in handle:
            line = clean_fasta_line(raw)
            if not line:
                continue
            if line.startswith(">"):
                flush_record()
                current_header = line
                continue
            if current_header is None:
                continue
            current_seq.append(line)

    flush_record()

    if not grouped:
        raise ValueError("No valid FASTA records found.")
    return grouped


def read_imgt_exon_lookup(fasta_path: Path) -> Dict[Tuple[str, str], str]:
    lookup: Dict[Tuple[str, str], str] = {}
    current_header: Optional[str] = None
    current_seq: List[str] = []

    def flush_record() -> None:
        nonlocal current_header, current_seq
        if not current_header:
            return
        seq = normalize_dna("".join(current_seq))
        if not seq:
            current_header = None
            current_seq = []
            return
        _, allele, exon, trim_5p, trim_3p = parse_imgt_header(current_header)
        if trim_5p > 0:
            seq = seq[trim_5p:] if trim_5p < len(seq) else ""
        if trim_3p > 0:
            seq = seq[:-trim_3p] if trim_3p < len(seq) else ""
        if seq:
            lookup[(allele, exon)] = seq
        current_header = None
        current_seq = []

    with fasta_path.open("r", encoding="utf-8", errors="ignore") as handle:
        for raw in handle:
            line = clean_fasta_line(raw)
            if not line:
                continue
            if line.startswith(">"):
                flush_record()
                current_header = line
                continue
            if current_header is not None:
                current_seq.append(line)
    flush_record()
    return lookup


def needleman_wunsch(ref: str, query: str, match: int = 2, mismatch: int = -1, gap: int = -2) -> Tuple[str, str]:
    n = len(ref)
    m = len(query)

    score = [[0] * (m + 1) for _ in range(n + 1)]
    trace = [[""] * (m + 1) for _ in range(n + 1)]

    for i in range(1, n + 1):
        score[i][0] = i * gap
        trace[i][0] = "U"
    for j in range(1, m + 1):
        score[0][j] = j * gap
        trace[0][j] = "L"

    for i in range(1, n + 1):
        ri = ref[i - 1]
        for j in range(1, m + 1):
            qj = query[j - 1]
            d = score[i - 1][j - 1] + (match if ri == qj else mismatch)
            u = score[i - 1][j] + gap
            l = score[i][j - 1] + gap
            best = max(d, u, l)
            score[i][j] = best
            if best == d:
                trace[i][j] = "D"
            elif best == u:
                trace[i][j] = "U"
            else:
                trace[i][j] = "L"

    i, j = n, m
    a_ref: List[str] = []
    a_query: List[str] = []
    while i > 0 or j > 0:
        t = trace[i][j] if i >= 0 and j >= 0 else ""
        if t == "D":
            a_ref.append(ref[i - 1])
            a_query.append(query[j - 1])
            i -= 1
            j -= 1
        elif t == "U":
            a_ref.append(ref[i - 1])
            a_query.append("-")
            i -= 1
        else:
            a_ref.append("-")
            a_query.append(query[j - 1])
            j -= 1

    return "".join(reversed(a_ref)), "".join(reversed(a_query))


def merge_ref_alignments(
    current_ref: str,
    new_ref: str,
    existing_rows: Dict[str, str],
    new_row: str,
) -> Tuple[str, Dict[str, str], str]:
    i = 0
    j = 0
    merged_ref: List[str] = []
    expanded_existing: Dict[str, List[str]] = {k: [] for k in existing_rows}
    expanded_new_row: List[str] = []

    while i < len(current_ref) or j < len(new_ref):
        a = current_ref[i] if i < len(current_ref) else None
        b = new_ref[j] if j < len(new_ref) else None

        consume_old = False
        consume_new = False

        if a is not None and b is not None and a == b:
            consume_old = True
            consume_new = True
            merged_ref.append(a)
        elif a == "-" and b != "-":
            consume_old = True
            merged_ref.append("-")
        elif b == "-" and a != "-":
            consume_new = True
            merged_ref.append("-")
        elif a is None and b is not None:
            consume_new = True
            merged_ref.append(b)
        elif b is None and a is not None:
            consume_old = True
            merged_ref.append(a)
        else:
            consume_old = True
            consume_new = True
            merged_ref.append(a if a is not None else (b if b is not None else "-"))

        for name, row in existing_rows.items():
            if consume_old and i < len(row):
                expanded_existing[name].append(row[i])
            else:
                expanded_existing[name].append("-")

        if consume_new and j < len(new_row):
            expanded_new_row.append(new_row[j])
        else:
            expanded_new_row.append("-")

        if consume_old:
            i += 1
        if consume_new:
            j += 1

    return "".join(merged_ref), {k: "".join(v) for k, v in expanded_existing.items()}, "".join(expanded_new_row)


def build_alignment(subtypes: Dict[str, str]) -> Dict[str, str]:
    names = sorted(subtypes.keys())
    ref_name = max(names, key=lambda n: len(subtypes[n]))
    ref_seq = subtypes[ref_name]

    msa_ref = ref_seq
    aligned_rows: Dict[str, str] = {ref_name: msa_ref}

    for name in names:
        if name == ref_name:
            continue
        a_ref, a_seq = needleman_wunsch(ref_seq, subtypes[name])
        msa_ref, aligned_rows, merged_new = merge_ref_alignments(msa_ref, a_ref, aligned_rows, a_seq)
        aligned_rows[name] = merged_new

    return aligned_rows


def find_shared_bounds(rows: Sequence[str]) -> Tuple[Optional[int], Optional[int]]:
    if not rows:
        return None, None
    length = len(rows[0])
    shared_start: Optional[int] = None
    shared_end: Optional[int] = None

    for i in range(length):
        if all(r[i] != "-" for r in rows):
            if shared_start is None:
                shared_start = i
            shared_end = i

    return shared_start, shared_end


def build_masked_consensus(rows: Sequence[str]) -> Tuple[str, Optional[int], Optional[int]]:
    if not rows:
        return "", None, None
    length = len(rows[0])
    for r in rows[1:]:
        if len(r) != length:
            raise ValueError("Alignment rows have inconsistent lengths.")

    shared_start, shared_end = find_shared_bounds(rows)
    if shared_start is None or shared_end is None:
        return "x" * length, None, None

    out: List[str] = []
    for i in range(length):
        col = [r[i] for r in rows]
        if i < shared_start or i > shared_end:
            out.append("x")
            continue

        if any(ch == "-" for ch in col):
            out.append("X")
            continue

        first = col[0]
        if all(ch == first for ch in col):
            out.append(first.lower())
        else:
            out.append("X")

    return "".join(out), shared_start, shared_end


def relation_to_consensus(consensus: str, subtype_row: str) -> Tuple[str, int]:
    out: List[str] = []
    mismatches = 0
    for c, s in zip(consensus, subtype_row):
        if s == "-":
            out.append("x")
            continue
        if c == "x":
            # Unshared edge in consensus: keep real base for this subtype.
            out.append(s.lower())
            continue
        if c == "X":
            # Variable position in consensus: show actual subtype base.
            out.append(s.upper())
            mismatches += 1
            continue
        if s == c:
            out.append(s.lower())
        else:
            out.append(s.upper())
            mismatches += 1
    return "".join(out), mismatches


def analyze_exon(gene: str, exon: str, subtypes: Dict[str, str]) -> ExonResult:
    aligned = build_alignment(subtypes)
    names = sorted(aligned.keys())
    rows = [aligned[n] for n in names]
    consensus, shared_start, shared_end = build_masked_consensus(rows)

    relation_rows: Dict[str, str] = {}
    mismatch_counts: Dict[str, int] = {}
    for name in names:
        rel, mm = relation_to_consensus(consensus, aligned[name])
        relation_rows[name] = rel
        mismatch_counts[name] = mm

    return ExonResult(
        gene=gene,
        exon=exon,
        consensus=consensus,
        subtype_rows=aligned,
        relation_rows=relation_rows,
        mismatch_counts=mismatch_counts,
        shared_start=shared_start,
        shared_end=shared_end,
    )


def safe_sheet_name(name: str, used: set[str]) -> str:
    cleaned = re.sub(r"[\\/*?:\[\]]", "_", name)
    cleaned = cleaned[:31] or "Sheet"
    if cleaned not in used:
        used.add(cleaned)
        return cleaned
    for i in range(2, 1000):
        cand = f"{cleaned[:28]}_{i}"
        if cand not in used:
            used.add(cand)
            return cand
    raise RuntimeError("Unable to allocate unique sheet name.")


def build_gene_consensus_summaries(
    results: List[ExonResult],
    annotations: Optional[List[JunctionAnnotation]] = None,
    family_mappings: Optional[Dict[str, set[str]]] = None,
) -> List[GeneConsensusSummary]:
    by_gene: Dict[str, Dict[str, ExonResult]] = defaultdict(dict)
    for res in results:
        by_gene[res.gene][res.exon] = res

    c_region_genes: set[str] = set()
    if annotations:
        c_region_genes = {
            ann.base_gene for ann in annotations if ann.original_region and ann.original_region.strip().upper() == "C-REGION"
        }

    out: List[GeneConsensusSummary] = []
    for gene in sorted(by_gene.keys()):
        exon_map = by_gene[gene]
        exon_labels = [lbl for lbl in exon_map.keys() if lbl.startswith("EX")]
        if exon_labels:
            by_family: Dict[str, List[str]] = defaultdict(list)
            active_families: Dict[str, set[str]] = {}
            if family_mappings:
                present = set(exon_labels)
                for fam, members in family_mappings.items():
                    present_members = present.intersection(members)
                    # Strict: collapse only if >1 family member is actually present in this gene.
                    if len(present_members) > 1:
                        active_families[fam] = present_members

            for lbl in exon_labels:
                fam_label = None
                for fam, members in active_families.items():
                    if lbl in members:
                        fam_label = fam
                        break
                key = fam_label if fam_label else lbl
                by_family[key].append(exon_map[lbl].consensus)
            ordered = sort_exon_family_labels(by_family.keys())
            parts = [merge_variant_consensus(by_family[lbl]) for lbl in ordered]
            seq = "".join(parts)
            pos = 1
            coords: List[str] = []
            lens: List[str] = []
            for lbl, part in zip(ordered, parts):
                end = pos + len(part) - 1
                coords.append(f"{lbl}:{pos}-{end}")
                lens.append(f"{lbl}:{len(part)}")
                pos = end + 1
            source = "C-REGION" if (gene in c_region_genes or "C-REGION" in exon_map) else "from_exons"
            out.append(
                GeneConsensusSummary(
                    gene=gene,
                    source=source,
                    exon_chain="+".join(ordered),
                    exon_relative_coords="+".join(coords),
                    exon_lengths="+".join(lens),
                    consensus=seq,
                )
            )
            continue

        ordered = sort_exon_labels(exon_map.keys())
        parts = [exon_map[lbl].consensus for lbl in ordered]
        seq = "".join(parts)
        pos = 1
        coords = []
        lens = []
        for lbl, part in zip(ordered, parts):
            end = pos + len(part) - 1
            coords.append(f"{lbl}:{pos}-{end}")
            lens.append(f"{lbl}:{len(part)}")
            pos = end + 1
        out.append(
            GeneConsensusSummary(
                gene=gene,
                source="C-REGION" if (gene in c_region_genes or "C-REGION" in exon_map) else "from_region",
                exon_chain="+".join(ordered),
                exon_relative_coords="+".join(coords),
                exon_lengths="+".join(lens),
                consensus=seq,
            )
        )
    return out


def find_shared_elements_between(
    gene_a: str,
    seq_a: str,
    gene_b: str,
    seq_b: str,
    min_len: int = 11,
    max_mismatches: int = 1,
) -> List[SharedElement]:
    n = len(seq_a)
    m = len(seq_b)
    if n == 0 or m == 0:
        return []

    elements: List[SharedElement] = []
    seen: set[Tuple[int, int, int, int]] = set()

    for delta in range(-(m - 1), n):
        i0 = max(0, delta)
        j0 = i0 - delta
        diag_len = min(n - i0, m - j0)
        if diag_len < min_len:
            continue

        left = 0
        mismatch_pos: List[int] = []
        for right in range(diag_len):
            ia = i0 + right
            jb = j0 + right
            if seq_a[ia] != seq_b[jb]:
                mismatch_pos.append(right)

            while len(mismatch_pos) > max_mismatches:
                left = mismatch_pos.pop(0) + 1

            win_len = right - left + 1
            if win_len < min_len:
                continue

            at_end = right == diag_len - 1
            if at_end:
                cannot_extend = True
            else:
                na = i0 + right + 1
                nb = j0 + right + 1
                next_is_mismatch = seq_a[na] != seq_b[nb]
                cannot_extend = next_is_mismatch and len(mismatch_pos) == max_mismatches

            if not cannot_extend:
                continue

            sa = i0 + left
            ea = i0 + right
            sb = j0 + left
            eb = j0 + right
            key = (sa, ea, sb, eb)
            if key in seen:
                continue
            seen.add(key)

            a_seg = seq_a[sa : ea + 1]
            b_seg = seq_b[sb : eb + 1]
            mm = sum(1 for x, y in zip(a_seg, b_seg) if x != y)
            elements.append(
                SharedElement(
                    gene_a=gene_a,
                    gene_b=gene_b,
                    start_a_1based=sa + 1,
                    end_a_1based=ea + 1,
                    start_b_1based=sb + 1,
                    end_b_1based=eb + 1,
                    length=len(a_seg),
                    mismatch_count=mm,
                    seq_a=a_seg,
                    seq_b=b_seg,
                )
            )

    elements.sort(key=lambda e: (-e.length, e.mismatch_count, e.gene_a, e.gene_b, e.start_a_1based, e.start_b_1based))
    return elements


def build_shared_elements_table(gene_summaries: List[GeneConsensusSummary]) -> List[SharedElement]:
    rows: List[SharedElement] = []
    for i in range(len(gene_summaries)):
        for j in range(i + 1, len(gene_summaries)):
            a = gene_summaries[i]
            b = gene_summaries[j]
            rows.extend(find_shared_elements_between(a.gene, a.consensus, b.gene, b.consensus, min_len=11, max_mismatches=1))
    rows.sort(key=lambda e: (-e.length, e.mismatch_count, e.gene_a, e.gene_b))
    return rows


def parse_compare_pairs(raw_pairs: Sequence[str]) -> List[Tuple[str, str]]:
    pairs: List[Tuple[str, str]] = []
    for raw in raw_pairs:
        txt = (raw or "").strip()
        if not txt:
            continue
        if "," in txt:
            a, b = [x.strip() for x in txt.split(",", 1)]
        elif ":" in txt:
            a, b = [x.strip() for x in txt.split(":", 1)]
        else:
            raise ValueError(f"Invalid --compare-pair '{raw}'. Use A,B")
        if not a or not b:
            raise ValueError(f"Invalid --compare-pair '{raw}'. Use A,B")
        pairs.append((a, b))
    return pairs


def build_pair_consensus_char(a: str, b: str) -> Tuple[str, str]:
    if a == "-" or b == "-":
        return "x", "x"
    if a == "X" or b == "X":
        return "X", "X"
    if a == "x" or b == "x":
        return "x", "x"
    au = a.upper()
    bu = b.upper()
    if au != bu:
        return "N", "N"
    return a.lower(), "|"


def build_pair_consensus_results(
    gene_summaries: List[GeneConsensusSummary],
    compare_pairs: Sequence[Tuple[str, str]],
) -> List[PairConsensusResult]:
    by_gene = {g.gene: g for g in gene_summaries}
    out: List[PairConsensusResult] = []
    for a, b in compare_pairs:
        ga = by_gene.get(a)
        gb = by_gene.get(b)
        if not ga or not gb:
            continue
        aa, bb = needleman_wunsch(ga.consensus, gb.consensus)
        cons_chars: List[str] = []
        mirrored_chars: List[str] = []
        for ca, cb in zip(aa, bb):
            c, m = build_pair_consensus_char(ca, cb)
            cons_chars.append(c)
            mirrored_chars.append(m)
        out.append(
            PairConsensusResult(
                gene_a=a,
                gene_b=b,
                aligned_a=aa,
                aligned_b=bb,
                mirrored="".join(mirrored_chars),
                consensus="".join(cons_chars),
            )
        )
    return out


def parse_family_mappings(family_args: Sequence[str]) -> Dict[str, set[str]]:
    # Default family mapping; can be extended/overridden by --family.
    mappings: Dict[str, set[str]] = {"EX2_family": {"EX2", "EX2R", "EX2T"}}
    for raw in family_args:
        text = (raw or "").strip()
        if not text or "=" not in text:
            raise ValueError(f"Invalid --family value '{raw}'. Expected format: FAMILY=MEM1,MEM2")
        fam, members_txt = text.split("=", 1)
        fam = fam.strip()
        members = {m.strip() for m in members_txt.split(",") if m.strip()}
        if not fam or not members:
            raise ValueError(f"Invalid --family value '{raw}'. Expected format: FAMILY=MEM1,MEM2")
        mappings[fam] = members
    return mappings


def build_effective_cli_call(
    input_path: Path,
    output_xlsx: Path,
    output_fasta: Path,
    exon_reference_fasta: Optional[Path],
    compare_pairs: Sequence[Tuple[str, str]],
    family_mappings: Dict[str, set[str]],
) -> str:
    parts: List[str] = [
        "consensus_cli.py",
        "--input",
        str(input_path),
        "--output",
        str(output_xlsx),
        "--consensus-fasta-output",
        str(output_fasta),
    ]
    if exon_reference_fasta:
        parts.extend(["--exon-reference-fasta", str(exon_reference_fasta)])
    for fam, members in sorted(family_mappings.items()):
        parts.extend(["--family", f"{fam}={','.join(sorted(members))}"])
    for a, b in compare_pairs:
        parts.extend(["--compare-pair", f"{a},{b}"])
    return " ".join(shlex.quote(p) for p in parts)


def capture_run_metadata(
    input_path: Path,
    output_xlsx: Path,
    output_fasta: Path,
    exon_reference_fasta: Optional[Path],
    cli_call_effective: str,
    compare_pairs: Sequence[Tuple[str, str]],
    family_mappings: Dict[str, set[str]],
) -> RunMetadata:
    def run_git(args: List[str]) -> str:
        try:
            out = subprocess.check_output(args, stderr=subprocess.DEVNULL, text=True).strip()
            return out
        except Exception:
            return ""

    git_commit = run_git(["git", "rev-parse", "HEAD"])
    git_branch = run_git(["git", "rev-parse", "--abbrev-ref", "HEAD"])
    git_remote_origin = run_git(["git", "remote", "get-url", "origin"])
    git_remotes = run_git(["git", "remote", "-v"])
    cli_call = " ".join(shlex.quote(a) for a in sys.argv)

    return RunMetadata(
        input_path=str(input_path),
        output_xlsx=str(output_xlsx),
        output_fasta=str(output_fasta),
        cwd=os.getcwd(),
        cli_call=cli_call,
        cli_call_effective=cli_call_effective,
        timestamp_utc=datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        python_version=sys.version.replace("\n", " "),
        platform=f"{platform.system()} {platform.release()} ({platform.machine()})",
        git_commit=git_commit,
        git_branch=git_branch,
        git_remote_origin=git_remote_origin,
        git_remotes=git_remotes.replace("\n", " ; "),
        exon_reference_fasta=str(exon_reference_fasta) if exon_reference_fasta else "",
        compare_pairs=";".join(f"{a},{b}" for a, b in compare_pairs),
        family_mappings=json.dumps({k: sorted(v) for k, v in sorted(family_mappings.items())}, sort_keys=True),
    )


def write_xlsx(
    results: List[ExonResult],
    out_path: Path,
    annotations: Optional[List[JunctionAnnotation]] = None,
    family_mappings: Optional[Dict[str, set[str]]] = None,
    compare_pairs: Optional[List[Tuple[str, str]]] = None,
    run_metadata: Optional[RunMetadata] = None,
) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except Exception as exc:
        raise RuntimeError("openpyxl is required for XLSX export. Install with: pip install openpyxl") from exc

    wb = Workbook()
    wb.remove(wb.active)
    used_names: set[str] = set()
    bold = Font(bold=True)

    gene_summaries = build_gene_consensus_summaries(results, annotations=annotations, family_mappings=family_mappings)
    pair_results = build_pair_consensus_results(gene_summaries, compare_pairs or [])

    gsummary = wb.create_sheet(title=safe_sheet_name("gene_summary", used_names))
    gsummary.append(
        [
            "gene",
            "source",
            "exon_chain",
            "exon_relative_coords_1based",
            "exon_lengths",
            "length",
            "X_count",
            "x_count",
            "consensus_fasta",
        ]
    )
    for gs in gene_summaries:
        gsummary.append(
            [
                gs.gene,
                gs.source,
                gs.exon_chain,
                gs.exon_relative_coords,
                gs.exon_lengths,
                len(gs.consensus),
                gs.consensus.count("X"),
                gs.consensus.count("x"),
                gs.consensus,
            ]
        )
    gsummary.append(["", ""])
    gsummary.append(["Legend", "X=polymorphism, x=padding"])
    gsummary.column_dimensions["A"].width = 14
    gsummary.column_dimensions["B"].width = 14
    gsummary.column_dimensions["C"].width = 28
    gsummary.column_dimensions["D"].width = 42
    gsummary.column_dimensions["E"].width = 30
    gsummary.column_dimensions["F"].width = 10
    gsummary.column_dimensions["G"].width = 10
    gsummary.column_dimensions["H"].width = 10
    gsummary.column_dimensions["I"].width = 140

    for pr in pair_results:
        pws = wb.create_sheet(title=safe_sheet_name(f"isotype_{pr.gene_a}_{pr.gene_b}", used_names))
        pws.append(["pair", f"{pr.gene_a} vs {pr.gene_b}"])
        pws.append([pr.gene_a, pr.aligned_a])
        pws.append([pr.gene_b, pr.aligned_b])
        pws.append(["mirrored_match", pr.mirrored])
        pws.append(["pair_consensus_fasta", pr.consensus])
        pws.append(["legend", "|=match, N=explicit mismatch, X=ambiguous (input X), x=gap/padding"])
        pws.column_dimensions["A"].width = 24
        pws.column_dimensions["B"].width = 160

    summary = wb.create_sheet(title=safe_sheet_name("summary", used_names))
    summary.append(
        [
            "gene",
            "exon",
            "isotype_count",
            "alignment_length",
            "shared_core_start_1based",
            "shared_core_end_1based",
            "X_count_in_consensus",
            "x_count_in_consensus",
            "consensus_fasta",
        ]
    )

    for res in results:
        aln_len = len(res.consensus)
        shared_start = (res.shared_start + 1) if res.shared_start is not None else ""
        shared_end = (res.shared_end + 1) if res.shared_end is not None else ""
        summary.append(
            [
                res.gene,
                res.exon,
                len(res.subtype_rows),
                aln_len,
                shared_start,
                shared_end,
                res.consensus.count("X"),
                res.consensus.count("x"),
                res.consensus,
            ]
        )
        ws_name = safe_sheet_name(f"{res.gene}_{res.exon}", used_names)
        ws = wb.create_sheet(title=ws_name)

        row = 1
        ws.cell(row=row, column=1, value=f"Gene: {res.gene}").font = bold
        ws.cell(row=row, column=2, value=f"Exon: {res.exon}").font = bold
        row += 1

        ws.cell(row=row, column=2, value="consensus_fasta").font = bold
        ws.cell(row=row, column=3, value=res.consensus)
        row += 1
        ws.cell(row=row, column=1, value="Legend").font = bold
        ws.cell(row=row, column=2, value="X=polymorphism, x=padding")
        row += 2

        ws.cell(row=row, column=1, value="isotype").font = bold
        ws.cell(row=row, column=2, value="single_base_changes_X_count").font = bold
        ws.cell(row=row, column=3, value="relation_to_consensus").font = bold
        row += 1

        for subtype in sorted(res.subtype_rows.keys()):
            ws.cell(row=row, column=1, value=subtype)
            ws.cell(row=row, column=2, value=res.mismatch_counts[subtype])
            ws.cell(row=row, column=3, value=res.relation_rows[subtype])
            row += 1

        ws.column_dimensions["A"].width = 24
        ws.column_dimensions["B"].width = 24
        ws.column_dimensions["C"].width = 140

    summary.append(["", ""])
    summary.append(["Legend", "X=polymorphism, x=padding"])

    summary.column_dimensions["A"].width = 14
    summary.column_dimensions["B"].width = 14
    summary.column_dimensions["C"].width = 14
    summary.column_dimensions["D"].width = 16
    summary.column_dimensions["E"].width = 22
    summary.column_dimensions["F"].width = 20
    summary.column_dimensions["G"].width = 16
    summary.column_dimensions["H"].width = 18
    summary.column_dimensions["I"].width = 120

    if annotations:
        jws = wb.create_sheet(title=safe_sheet_name("junctions", used_names))
        jws.append(
            [
                "allele",
                "base_gene",
                "original_region",
                "split_source",
                "exon_chain",
                "junction_boundaries_1based",
                "note",
            ]
        )
        for ann in annotations:
            jws.append(
                [
                    ann.allele,
                    ann.base_gene,
                    ann.original_region,
                    ann.split_source,
                    ann.exon_chain,
                    ann.junction_boundaries_1based,
                    ann.note,
                ]
            )
        jws.column_dimensions["A"].width = 18
        jws.column_dimensions["B"].width = 12
        jws.column_dimensions["C"].width = 22
        jws.column_dimensions["D"].width = 24
        jws.column_dimensions["E"].width = 22
        jws.column_dimensions["F"].width = 28
        jws.column_dimensions["G"].width = 44

    if run_metadata:
        rws = wb.create_sheet(title=safe_sheet_name("reproducibility", used_names))
        rws.append(["field", "value"])
        rows = [
            ("timestamp_utc", run_metadata.timestamp_utc),
            ("cli_call_raw", run_metadata.cli_call),
            ("cli_call_effective", run_metadata.cli_call_effective),
            ("cwd", run_metadata.cwd),
            ("input_path", run_metadata.input_path),
            ("output_xlsx", run_metadata.output_xlsx),
            ("output_fasta", run_metadata.output_fasta),
            ("exon_reference_fasta", run_metadata.exon_reference_fasta),
            ("compare_pairs", run_metadata.compare_pairs),
            ("family_mappings", run_metadata.family_mappings),
            ("python_version", run_metadata.python_version),
            ("platform", run_metadata.platform),
            ("git_commit", run_metadata.git_commit),
            ("git_branch", run_metadata.git_branch),
            ("git_remote_origin", run_metadata.git_remote_origin),
            ("git_remotes", run_metadata.git_remotes),
        ]
        for k, v in rows:
            rws.append([k, v])
        rws.column_dimensions["A"].width = 28
        rws.column_dimensions["B"].width = 160

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def wrap_sequence(seq: str, width: int = 80) -> Iterable[str]:
    for i in range(0, len(seq), width):
        yield seq[i : i + width]


def write_consensus_fasta(
    results: List[ExonResult],
    out_path: Path,
    pair_results: Optional[List[PairConsensusResult]] = None,
) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", encoding="utf-8") as handle:
        for res in results:
            handle.write(f">{res.gene}|{res.exon}|consensus|n={len(res.subtype_rows)}\n")
            for line in wrap_sequence(res.consensus, width=80):
                handle.write(line + "\n")
        for pr in pair_results or []:
            handle.write(f">{pr.gene_a}_vs_{pr.gene_b}|isotype_pair_consensus\n")
            for line in wrap_sequence(pr.consensus, width=80):
                handle.write(line + "\n")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Compute per-gene, per-exon consensus from IMGT FASTA or compiled CSV.")
    parser.add_argument("--input", required=True, help="Input IMGT FASTA or compiled CSV path")
    parser.add_argument("--output", required=True, help="Output XLSX path")
    parser.add_argument(
        "--exon-reference-fasta",
        default="",
        help="Optional IMGT exon FASTA used to split CSV rows when coordinates are missing.",
    )
    parser.add_argument(
        "--consensus-fasta-output",
        default="",
        help="Optional output FASTA path (default: <output_stem>_consensus.fasta)",
    )
    parser.add_argument(
        "--family",
        action="append",
        default=[],
        help="Family collapse mapping for gene_summary (repeatable): FAMILY=MEM1,MEM2. "
        "Collapsed only when >1 family member is present in that gene.",
    )
    parser.add_argument(
        "--compare-pair",
        action="append",
        default=[],
        help="Explicit pair comparison for isotype pages/FASTA (repeatable): GENE_A,GENE_B. "
        "Default includes TRGC1,TRGC2 and TRBC1,TRBC2.",
    )
    parser.add_argument("--gene", action="append", default=[], help="Restrict to one or more base genes (repeatable)")
    parser.add_argument("--exon", action="append", default=[], help="Restrict to one or more exon labels (repeatable)")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    inp = Path(args.input).expanduser().resolve()
    out_xlsx = Path(args.output).expanduser().resolve()
    family_mappings = parse_family_mappings(args.family)
    compare_pairs = parse_compare_pairs(args.compare_pair or ["TRGC1,TRGC2", "TRBC1,TRBC2"])

    annotations: List[JunctionAnnotation] = []
    if inp.suffix.lower() == ".csv":
        exon_lookup = None
        used_exon_ref: Optional[Path] = None
        if args.exon_reference_fasta:
            ref_path = Path(args.exon_reference_fasta).expanduser().resolve()
            if ref_path.exists():
                exon_lookup = read_imgt_exon_lookup(ref_path)
                used_exon_ref = ref_path
        else:
            sibling_ref = inp.parent / "TCRC_exons.fasta"
            if sibling_ref.exists():
                exon_lookup = read_imgt_exon_lookup(sibling_ref)
                used_exon_ref = sibling_ref
        grouped, annotations = read_compiled_csv_sequences(inp, exon_lookup=exon_lookup)
    else:
        used_exon_ref = None
        grouped = read_imgt_fasta_sequences(inp)
    target_genes = set(args.gene) if args.gene else None
    target_exons = set(args.exon) if args.exon else None

    results: List[ExonResult] = []
    for gene, exon in sorted(grouped.keys()):
        if target_genes is not None and gene not in target_genes:
            continue
        if target_exons is not None and exon not in target_exons:
            continue
        results.append(analyze_exon(gene, exon, grouped[(gene, exon)]))

    if not results:
        raise SystemExit("No gene/exon groups matched the input/filter settings.")

    if args.consensus_fasta_output:
        out_fasta = Path(args.consensus_fasta_output).expanduser().resolve()
    else:
        out_fasta = out_xlsx.with_name(f"{out_xlsx.stem}_consensus.fasta")
    cli_call_effective = build_effective_cli_call(
        input_path=inp,
        output_xlsx=out_xlsx,
        output_fasta=out_fasta,
        exon_reference_fasta=used_exon_ref,
        compare_pairs=compare_pairs,
        family_mappings=family_mappings,
    )
    run_metadata = capture_run_metadata(
        input_path=inp,
        output_xlsx=out_xlsx,
        output_fasta=out_fasta,
        exon_reference_fasta=used_exon_ref,
        cli_call_effective=cli_call_effective,
        compare_pairs=compare_pairs,
        family_mappings=family_mappings,
    )
    gene_summaries = build_gene_consensus_summaries(results, annotations=annotations, family_mappings=family_mappings)
    pair_results = build_pair_consensus_results(gene_summaries, compare_pairs)
    write_xlsx(
        results,
        out_xlsx,
        annotations=annotations,
        family_mappings=family_mappings,
        compare_pairs=compare_pairs,
        run_metadata=run_metadata,
    )
    print(f"Wrote XLSX: {out_xlsx}")
    write_consensus_fasta(results, out_fasta, pair_results=pair_results)
    print(f"Wrote consensus FASTA: {out_fasta}")


if __name__ == "__main__":
    main()
